import openpyxl
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import ExpressionWrapper
from django.http import HttpResponse
from django.views.generic import View, TemplateView
from app.ventas.models import Producto, Local, Asistencia, Trabajador, GastosDiario
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect
from openpyxl.styles import Font, PatternFill, Border, Side, numbers
from django.contrib import messages
from django.shortcuts import render, redirect
from app.ventas.forms import VentaDiariaForm
from django.views.decorators.csrf import csrf_exempt
from .models import VentaDiaria
from django.shortcuts import render
from django.http import JsonResponse
from datetime import datetime
from django.shortcuts import get_object_or_404
from django.db.models import ExpressionWrapper, Sum, F, FloatField
from django.db.models.functions import Coalesce
from decimal import Decimal

class Home(LoginRequiredMixin,TemplateView):
    template_name = 'admin/ventas_por_local.html'


class DescargarDatosExcelNomina(LoginRequiredMixin,View):
    # Crear un nuevo libro de Excel
    def get(self, request, *args, **kwargs):
        # datos_por_fecha = request.GET.get('datos_por_fecha')
        local_id = request.GET.get('local_id')
        local = Local.objects.get(id=local_id)
        # obtener las fechas de los últimos 7 días
        hoy = datetime.today().date()
        # fechas = [hoy - timedelta(days=i) for i in range(7)]
        ventas = VentaDiaria.objects.filter(local=local)
        fechas = ventas.dates('fecha', 'day', order='DESC')
        # recopilar los datos de ventas diarias, asistencias, y totales de importe para cada fecha y local
        datos_por_fecha = {}
        cant_trabajadores = {}
        trabajadores_salarios = []
        ganancia_fecha = {}
        asistencias = {}
        trabajadores = Trabajador.objects.filter(local=local)

        for fecha in fechas:
            datos_por_fecha[fecha] = {}
            venta_diaria = VentaDiaria.objects.filter(fecha=fecha, local=local).select_related('local', 'producto')
            total_importe_ventas = venta_diaria.aggregate(total_ventas=
                                                          Sum(F('cantidad_venta') * F('precio_venta_producto')))[
                'total_ventas']
            importe_venta_elaborados = venta_diaria.filter(producto__categoria='1').aggregate(total_ventas=
                                                                                              Sum(F(
                                                                                                  'cantidad_venta') * F(
                                                                                                  'precio_venta_producto')))[
                'total_ventas']

            importe_venta_no_elaborados = venta_diaria.filter(producto__categoria='2').aggregate(total_ventas=
                                                                                                 Sum(F(
                                                                                                     'cantidad_venta') * F(
                                                                                                     'precio_venta_producto')))[
                'total_ventas']
            ganancia = total_importe_ventas

            cantidad_trabajadores_dia = Asistencia.objects.filter(fecha=fecha, local=local).select_related(
                'trabajador').distinct().count()

            if ganancia > 15000:
                ganancia -= 15000

                ganancia = (importe_venta_elaborados * Decimal(str('0.06'))) + (
                            (ganancia - importe_venta_elaborados) * Decimal(str('0.02')))
                ganancia /= cantidad_trabajadores_dia

            else:
                ganancia = 0
            ganancia_fecha[fecha] = ganancia

            asistencias = Asistencia.objects.filter(fecha=fecha, local=local).select_related('trabajador').annotate(
                salario_devengado_diario=ExpressionWrapper(F('trabajador__salario_basico') + ganancia,
                                                           output_field=FloatField()),
            )

            # trabajadores_salarios


            # Creamos la tupla con el queryset y los otros valores
            # ventas_por_fecha[fecha] = (ventas_fecha, total_importe_venta, total_importe_costo, ganancia_dia)
            datos_por_fecha[fecha] = (
            venta_diaria, asistencias, ganancia, importe_venta_elaborados, importe_venta_no_elaborados)
            cant_trabajadores = Trabajador.objects.filter(local=local).count()

        a= asistencias
        # determinar remuneracion por trabajador
        for trabajador in trabajadores:
            remuneracion_total = Decimal(str('0'))
            for _, d in datos_por_fecha.items():
                for a in d[1].filter(trabajador=trabajador):
                    remuneracion_total += Decimal(str(a.salario_devengado_diario))
            trabajadores_salarios.append({
                'nombre': trabajador.nombre,
                'total_devengado': remuneracion_total
            },)




        filename = local.nombre
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Resumen datos"

        # Escribir los encabezados de la tabla en la hoja de Excel
        sheet.cell(row=1, column=1, value="Fecha")
        sheet.cell(row=1, column=2, value="Trabajador")
        sheet.cell(row=1, column=3, value="Salario Básico")
        sheet.cell(row=1, column=4, value="Salario Devengado al Día")
        sheet.cell(row=1, column=5, value="Ganancia por Ventas al Día")

        # Escribir los datos de la tabla en hojas de Excel separadas por fecha
        for fecha, data in datos_por_fecha.items():

            # Crear una nueva hoja de Excel para cada fecha
            sheet = workbook.create_sheet(title=fecha.strftime("%Y-%m-%d"))

            # Escribir los encabezados de la tabla en la hoja de Excel
            sheet.cell(row=1, column=1, value="Fecha")
            sheet.cell(row=1, column=2, value="Trabajador")
            sheet.cell(row=1, column=3, value="Salario Básico")
            sheet.cell(row=1, column=4, value="Salario Devengado al Día")
            sheet.cell(row=1, column=5, value="Ganancia por Ventas al Día")

            # Escribir los datos de la tabla en la hoja de Excel para cada fecha
            for index, row_data in enumerate(data[1]):
                sheet.cell(row=index + 2, column=1, value=row_data.fecha)
                sheet.cell(row=index + 2, column=2, value=row_data.trabajador.nombre)
                sheet.cell(row=index + 2, column=3, value=row_data.trabajador.salario_basico)
                sheet.cell(row=index + 2, column=4, value=row_data.salario_devengado_diario)
                sheet.cell(row=index + 2, column=5, value=data[2])

        # Crear una nueva hoja de Excel para el resumen
        sheet = workbook.create_sheet(title="Resumen")

        # Escribir los encabezados de la tabla en la hoja de Excel
        sheet.cell(row=1, column=1, value="Trabajador")
        sheet.cell(row=1, column=2, value="Salario a pagar")

        # Escribir los datos del resumen en la hoja de Excel
        index = 0
        for row_data in trabajadores_salarios:
            index += 1
            sheet.cell(row=index + 2, column=1, value=row_data['nombre'])
            sheet.cell(row=index + 2, column=2, value=row_data['total_devengado'])

            # Eliminar la hoja de Excel por defecto

            # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + filename + 'Nomina.xlsx'
        workbook.save(response)
        return response


class DescargarExcelVentasDiariasView(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        # Obtener el local específico
        local_id = request.GET.get('local_id')
        local_obj = Local.objects.get(id=local_id)

        # Obtener las fechas de venta únicas para el local
        fechas = VentaDiaria.objects.filter(local=local_obj).order_by('fecha').values_list('fecha', flat=True).distinct()
        sum_total_importe_precio_venta = 0
        filename = ""

        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()

        # Crear una hoja de Excel para cada fecha de venta
        for fecha in fechas:
            # Obtener las ventas para la fecha actual y el local específico
            ventas = VentaDiaria.objects.filter(local=local_obj, fecha=fecha).all()

            # Crear una nueva hoja de Excel para la fecha actual
            ws = wb.create_sheet(str(fecha))

            # Definir el estilo para los encabezados de tabla
            bold_font = Font(bold=True)
            bg_fill = PatternFill(fill_type='solid', start_color='CCE5FF', end_color='CCE5FF')
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Escribir los encabezados de la tabla en la primera fila
            ws['A1'] = "Producto"
            ws['B1'] = "Cantidad vendida"
            ws['C1'] = "Precio de venta"
            ws['D1'] = "Importe de venta"
            ws['E1'] = "Precio de costo"
            ws['F1'] = "Importe de costo"

            # Aplicar el estilo a los encabezados de tabla
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws[f'{col}1']
                cell.font = bold_font
                cell.fill = bg_fill
                cell.border = border

            # Escribir los datos de la tabla en las filas siguientes
            row_num = 2
            sum_importe_precio_costo = 0
            sum_importe_precio_venta = 0
            sum_importe_precio_costo_t = sum(
                s.importe_precio_costo for s in VentaDiaria.objects.filter(local=local_obj, fecha__lte=fecha))
            sum_importe_precio_venta_t = sum(
                s.importe_precio_venta for s in VentaDiaria.objects.filter(local=local_obj, fecha__lte=fecha))
            sum_total_importe_precio_venta = sum_importe_precio_venta_t - sum_importe_precio_costo_t
            for venta in ventas:
                filename = f"{venta.local.nombre}-{str(venta.fecha)}"
                ws.cell(row=row_num, column=1, value=str(venta.producto))
                ws.cell(row=row_num, column=2, value=venta.cantidad_venta)
                ws.cell(row=row_num, column=3, value=venta.precio_venta_producto)
                ws.cell(row=row_num, column=4, value=venta.importe_precio_venta)
                sum_importe_precio_venta += venta.importe_precio_venta
                ws.cell(row=row_num, column=5, value=venta.precio_costo_producto)
                ws.cell(row=row_num, column=6, value=venta.importe_precio_costo)
                sum_importe_precio_costo += venta.importe_precio_costo
                row_num += 1
            ws.cell(row=row_num, column=1, value="Total Importe por ventas")

            ws.cell(row=row_num, column=2, value=str(sum_importe_precio_venta))
            row_num += 1
            ws.cell(row=row_num, column=1, value="Total importe por precio costo")

            ws.cell(row=row_num, column=2, value=str(sum_importe_precio_costo))
            row_num += 1
            ws.cell(row=row_num, column=1, value="Ganancia en el dia")

            ws.cell(row=row_num, column=2, value=str(sum_importe_precio_venta - sum_importe_precio_costo))

            row_num += 1
            ws.cell(row=row_num, column=1, value="Ganancia Total")

            ws.cell(row=row_num, column=2, value=str(sum_total_importe_precio_venta))

            # Formatear las celdas de importe de venta y costo como moneda
            currency_format = '#,##0.00 $'
            for row in ws.iter_rows(min_row=2, max_row=row_num, min_col=4, max_col=6):
                for cell in row:
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            # Ajustar el ancho de las columnas automáticamente
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # Eliminar la hoja de Excel por defecto
        wb.remove(wb['Sheet'])

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + filename + '.xlsx'
        wb.save(response)

        return response

# @login_required
# def importar_ventas(request):
#     if request.method == 'POST' and request.FILES['archivo']:
#         archivo = request.FILES['archivo']
#
#         try:
#             # Leer el contenido del archivo Excel y convertirlo en un DataFrame de pandas
#             df = pd.read_excel(archivo)
#
#             # Crear una instancia de VentaDiariaForm para validar los datos
#             form = VentaDiariaForm()
#
#             # Iterar sobre las filas del DataFrame y crear instancias de VentaDiaria
#             for i, fila in df.iterrows():
#                 data = {
#                     'local': fila['Local'],
#                     'producto': fila['Producto'],
#                     'cantidad_entrada': fila['Cantidad Entrada'],
#                     'cantidad_merma': fila['Cantidad Merma'],
#                     'fecha': fila['Fecha'],
#                     'cantidad_venta': fila['Cantidad Venta'],
#                 }
#                 form = VentaDiariaForm(data)
#                 if form.is_valid():
#                     form.save()
#                 else:
#                     messages.warning(request, f'Error en la fila {i + 1}: {form.errors}')
#
#             messages.success(request, 'Las ventas se han importado correctamente.')
#             return redirect('admin:ventas_ventadiaria_changelist')
#         except Exception as e:
#             messages.error(request, f'Error al importar el archivo: {e}')
#     else:
#         form = VentaDiariaForm()
#
#     context = {
#         'form': form,
#     }
#     return render(request, 'ventas/importar_ventas.html', context)
#

@login_required
@csrf_protect
@ensure_csrf_cookie
def eliminar_venta(request):
    if request.method == 'POST':
        venta_id = request.POST.get('id')
        venta = VentaDiaria.objects.get(id=venta_id)
        if venta:
            venta.delete()
            return JsonResponse({'mensaje': 'Venta eliminada exitosamente.'})
        return JsonResponse({'error': 'Venta no eliminada exitosamente.'})

@login_required
@csrf_protect
@ensure_csrf_cookie
def eliminar_gasto(request):
    if request.method == 'POST':
        gasto_id = request.POST.get('id')
        gasto = GastosDiario.objects.get(id=gasto_id)
        if gasto:
            gasto.delete()
            return JsonResponse({'mensaje': 'Gasto  eliminada exitosamente.'})
        return JsonResponse({'error': 'Gasto  no eliminada exitosamente.'})



@login_required
@csrf_exempt
def actualizar_venta_diaria(request):
    if request.method == 'POST':
        venta_id = request.POST.get('id')
        nueva_cantidad_venta = request.POST.get('cantidad_venta')
        nueva_cantidad_merma = request.POST.get('cantidad_merma')

        venta_diaria = VentaDiaria.objects.get(id=venta_id)
        venta_diaria.cantidad_venta = nueva_cantidad_venta
        venta_diaria.cantidad_merma = nueva_cantidad_merma
        venta_diaria.save()

        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Método de solicitud no válido.'})

@login_required
def buscar_productos(request):
    query = request.GET.get('q')
    productos = Producto.objects.filter(nombre__icontains=query)
    results = []
    for producto in productos:
        results.append({
            'id': producto.id,
            'text': producto.nombre
        })
    return JsonResponse({'results': results})


@login_required
@csrf_protect
@ensure_csrf_cookie
def eliminar_ventas(request):
    if request.method == 'POST':
        fecha_datetime = datetime.strptime(request.POST.get('fecha'), '%B %d, %Y')
        fecha_formateada = fecha_datetime.strftime('%Y-%m-%d')
        local = request.POST.get('local')

        ventas = VentaDiaria.objects.filter(fecha=fecha_formateada, local_id=local)
        total_eliminadas, _ = ventas.delete()
        return JsonResponse({'total_eliminadas': total_eliminadas})
    else:
        return JsonResponse({'error': 'Método no permitido'})


@login_required
def ventas_diarias(request):
    fecha_obj = datetime.strptime(request.GET.get('fecha'), '%d/%m/%Y')
    ventas = VentaDiaria.objects.filter(fecha=fecha_obj, local_id=request.GET.get('local_id'))
    print(request.GET.get('local_id'))
    # Aquí obtienes las ventas diarias del local en la fecha especificada
    data = []
    for venta in ventas:
        data.append({
            'id': venta.id,
            'producto': venta.producto.nombre,
            'fecha': venta.fecha.strftime('%d de %B de %Y'),
            'cantidad_disponible': venta.cantidad_disponible,
            'cantidad_merma': venta.cantidad_merma,
            'cantidad_venta': venta.cantidad_venta,
            'precio_venta_producto': venta.precio_venta_producto,
            'importe_precio_venta': venta.importe_precio_venta,
            'precio_costo_producto': venta.precio_costo_producto,
            'importe_precio_costo': venta.importe_precio_costo,
        })
    return JsonResponse({'data': data})



@login_required
@csrf_protect
@ensure_csrf_cookie
def eliminar_asistencia(request):
    if request.method == 'POST':
        asistencia_id = request.POST.get('id')
        asistencia = Asistencia.objects.get(id=asistencia_id)
        if asistencia:
            asistencia.delete()
            return JsonResponse({'mensaje': 'asistencia eliminada exitosamente.'})
        return JsonResponse({'error': 'asistencia no eliminada exitosamente.'})



@login_required
@csrf_protect
@csrf_exempt
@ensure_csrf_cookie
def eliminar_trabajador(request):
    if request.method == 'POST':
        trabajador_id = request.POST.get('id')
        trabajador = get_object_or_404(Trabajador, pk=trabajador_id)
        if trabajador:
            trabajador.delete()
            return JsonResponse({'mensaje': 'trabajador eliminada exitosamente.'})
        return JsonResponse({'error': 'trabajador no eliminada exitosamente.'})


@login_required
@csrf_protect
@csrf_exempt
@ensure_csrf_cookie
def actualizar_trabajador(request):
    if request.method == 'POST':
        trabajador_id = request.POST.get('id')
        nombre = request.POST.get('nombre')
        salario_basico = request.POST.get('salario_basico')

        trabajador = Trabajador.objects.get(id=trabajador_id)
        trabajador.nombre = nombre
        trabajador.salario_basico = salario_basico
        trabajador.save()

        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Método de solicitud no válido.'})

